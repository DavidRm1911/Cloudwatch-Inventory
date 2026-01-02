"""
AWS CloudWatch Alarms Multi-Account Inventory Export

This Lambda function collects CloudWatch alarms from multiple AWS accounts across
specified regions and exports them to an Excel file stored in S3.

Purpose:
- Centralized monitoring inventory across AWS Organization
- Cross-account CloudWatch alarm auditing and compliance reporting
- Automated alarm documentation and tracking

Use Case:
- Cloud Operations teams managing multiple AWS accounts
- Security and compliance auditing
- Cost optimization and alarm coverage analysis

Architecture:
- Assumes cross-account IAM role in each target account
- Collects alarms from CloudWatch API
- Generates formatted Excel workbook with one sheet per account
- Uploads final report to S3 bucket

Author: Cloud Operations Team
"""

import boto3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


def lambda_handler(event, context):
    """
    Main Lambda handler function
    
    Process:
    1. Iterate through all configured AWS accounts
    2. Assume cross-account role in each account (or use default credentials for same account)
    3. Query CloudWatch alarms in specified regions
    4. Extract alarm details, metrics, and tags
    5. Generate Excel workbook with formatted data
    6. Upload to S3 for distribution
    """
    
    # Initialize AWS clients
    sts_client = boto3.client('sts')
    s3_client = boto3.client('s3')

    # Get Lambda's current account ID to avoid unnecessary role assumption
    current_account_id = sts_client.get_caller_identity()['Account']

    # Configuration
    ROLE_NAME = "CrossAccountCloudWatchReadRole"
    REGIONS = ['us-east-1', 'us-west-2']
    
    account_ids = {
        '111111111111': 'production-account',
        '222222222222': 'staging-account',
        '333333333333': 'development-account',
        '444444444444': 'security-account',
        '555555555555': 'shared-services',
    }

    # Excel workbook and styling setup
    workbook = openpyxl.Workbook()
    first_sheet = True

    header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    base_headers = [
        'Account', 'Region', 'Service', 'AlarmName', 'MetricName', 'Statistic',
        'Period', 'Threshold', 'Datapoints', 'Actions', 'ResourceId', 'Description', 'State'
    ]
    
    tag_keys_set = set()
    alarms_by_account = {}

    # Process each AWS account
    for account_id, account_name in account_ids.items():
        try:
            alarms_list = []
            
            # Process each region
            for region in REGIONS:
                try:
                    # Create CloudWatch client with appropriate credentials
                    if account_id == current_account_id:
                        cloudwatch_client = boto3.client('cloudwatch', region_name=region)
                    else:
                        assumed_role = sts_client.assume_role(
                            RoleArn=f"arn:aws:iam::{account_id}:role/{ROLE_NAME}",
                            RoleSessionName="AlarmInventorySession"
                        )
                        credentials = assumed_role['Credentials']
                        
                        cloudwatch_client = boto3.client(
                            'cloudwatch',
                            region_name=region,
                            aws_access_key_id=credentials['AccessKeyId'],
                            aws_secret_access_key=credentials['SecretAccessKey'],
                            aws_session_token=credentials['SessionToken']
                        )

                    # Query CloudWatch alarms using pagination
                    paginator = cloudwatch_client.get_paginator('describe_alarms')
                    
                    for page in paginator.paginate():
                        for alarm in page['MetricAlarms']:
                            try:
                                # Parse alarm namespace and metric
                                namespace = alarm.get('Namespace', 'Custom')
                                if namespace and namespace.startswith('AWS/'):
                                    namespace = namespace.replace('AWS/', '')
                                metric_name = alarm.get('MetricName', 'N/A')

                                # Parse threshold and evaluation configuration
                                comparison_operator = alarm.get('ComparisonOperator', 'N/A')
                                threshold = alarm.get('Threshold', 'N/A')
                                evaluation_periods = alarm.get('EvaluationPeriods', 1)
                                period = alarm.get('Period', 300)
                                datapoints_to_alarm = alarm.get('DatapointsToAlarm')

                                threshold_text = f"{metric_name} {comparison_operator} {threshold} for {evaluation_periods} datapoints within {period//60} minutes"
                                period_text = f"{period//60} minutes"
                                datapoint_text = f"{datapoints_to_alarm} out of {evaluation_periods}" if datapoints_to_alarm else f"{evaluation_periods} out of {evaluation_periods}"

                                # Extract resource ID from alarm dimensions
                                resource_id = ''
                                id_type = ''
                                priority_dimensions = [
                                    'InstanceId', 'DBInstanceIdentifier', 'FunctionName',
                                    'LoadBalancerName', 'TargetGroup', 'ClusterName',
                                    'WebACL', 'BucketName', 'QueueName', 'TopicName',
                                    'AutoScalingGroupName', 'CacheClusterId', 'TableName'
                                ]
                                
                                for dim in alarm.get('Dimensions', []):
                                    dim_name = dim['Name']
                                    dim_value = dim['Value']

                                    if dim_name in priority_dimensions:
                                        resource_id = dim_value
                                        id_type = dim_name
                                        break
                                    elif not resource_id:
                                        resource_id = dim_value
                                        id_type = dim_name

                                formatted_resource_id = f"{resource_id} ({id_type})" if id_type and resource_id else resource_id

                                # Parse and format alarm actions
                                actions_list = []
                                for arn in alarm.get('AlarmActions', []):
                                    if arn.startswith('arn:aws:autoscaling'):
                                        actions_list.append(f'Execute autoscaling action: {arn}')
                                    elif arn.startswith('arn:aws:sns'):
                                        topic_name = arn.split(':')[-1]
                                        actions_list.append(f'Send notification to SNS topic: {topic_name}')
                                    elif arn.startswith('arn:aws:ec2'):
                                        actions_list.append(f'Reboot EC2 instance: {resource_id}')
                                    else:
                                        actions_list.append(f'Execute action: {arn}')

                                actions_str = '\n'.join(actions_list) if actions_list else 'No actions configured'
                                description = alarm.get('AlarmDescription', 'No description')
                                
                                # Skip AWS-managed alarms
                                if "DO NOT EDIT OR DELETE" in description.upper():
                                    continue

                                # Retrieve alarm tags
                                try:
                                    tags_response = cloudwatch_client.list_tags_for_resource(
                                        ResourceARN=alarm['AlarmArn']
                                    )
                                    tags = tags_response.get('Tags', [])
                                    tag_dict = {tag['Key']: tag['Value'] for tag in tags}
                                    tag_keys_set.update(tag_dict.keys())
                                except Exception:
                                    tag_dict = {}

                                # Collect alarm data
                                alarms_list.append({
                                    'Account': account_name,
                                    'Region': region,
                                    'Service': namespace,
                                    'AlarmName': alarm.get('AlarmName', 'N/A'),
                                    'MetricName': metric_name,
                                    'Statistic': alarm.get('Statistic', alarm.get('ExtendedStatistic', 'N/A')),
                                    'Period': period_text,
                                    'Threshold': threshold_text,
                                    'Datapoints': datapoint_text,
                                    'Actions': actions_str,
                                    'ResourceId': formatted_resource_id,
                                    'Description': description,
                                    'State': alarm.get('StateValue', 'UNKNOWN'),
                                    'Tags': tag_dict
                                })

                            except Exception as alarm_error:
                                # Log alarm-level errors but continue with other alarms
                                print(f"Error processing alarm in {region}/{account_id}: {str(alarm_error)}")
                                continue

                except Exception as region_error:
                    # Log region-level errors but continue with other regions
                    print(f"Error processing region {region} in account {account_id}: {str(region_error)}")
                    continue

            # Store alarms for this account
            alarms_by_account[account_name] = alarms_list

        except Exception as account_error:
            # Log account-level errors but continue with other accounts
            print(f"Error processing account {account_id}: {str(account_error)}")
            continue

    # Build Excel workbook with collected alarm data
    tag_headers = sorted(tag_keys_set)
    all_headers = base_headers + tag_headers

    for account_name, alarms_list in alarms_by_account.items():
        # Create worksheet for each account
        sheet = workbook.create_sheet(title=account_name) if not first_sheet else workbook.active
        if first_sheet:
            sheet.title = account_name
            first_sheet = False

        # Write and style header row
        sheet.append(all_headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill

        # Write alarm data sorted by service
        alarms_list.sort(key=lambda x: x['Service'].lower())

        for alarm in alarms_list:
            row_data = [
                alarm['Account'], alarm['Region'], alarm['Service'], alarm['AlarmName'],
                alarm['MetricName'], alarm['Statistic'], alarm['Period'], alarm['Threshold'],
                alarm['Datapoints'], alarm['Actions'], alarm['ResourceId'], alarm['Description'],
                alarm['State']
            ]
            for key in tag_headers:
                row_data.append(alarm['Tags'].get(key, ''))
            sheet.append(row_data)

        # Apply styling to data cells
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(all_headers)):
            for cell in row:
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 5

    # Export workbook to S3
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Upload to S3 bucket
    bucket_name = 'cloudwatch-inventory-reports'
    file_name = 'cloudwatch-alarms-inventory.xlsx'
    
    try:
        s3_client.put_object(
            Bucket=bucket_name,
            Key=file_name,
            Body=excel_file.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        print(f"Successfully uploaded {file_name} to {bucket_name}")
        return {
            'statusCode': 200,
            'body': f'Successfully generated and uploaded CloudWatch alarms inventory to s3://{bucket_name}/{file_name}'
        }
    except Exception as s3_error:
        print(f"Error uploading to S3: {str(s3_error)}")
        return {
            'statusCode': 500,
            'body': f'Error uploading to S3: {str(s3_error)}'
        }
